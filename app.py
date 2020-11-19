from flask import Flask, render_template, request
from openpyxl import load_workbook

app = Flask(__name__)


@app.route("/")
def index():
     
     #мы делим строку на [описание, ссылка] картинки
     images = []
     excel = load_workbook("gallery.xlsx")
     page = excel["Лист1"]
     for row in page:
          url = row[0].value
          description = row[1].value
          title = row[2].value
          lst = [title, description, url]
          images.append(lst)

     return render_template("index.html", images=images)


@app.route("/add")
def add():
     return render_template ("form.html")


@app.route("/reciever", methods=["POST"])
def reciever():
     description = request.form.get("description")
     url = request.form.get("url")
     title = request.form.get("title")

     excel = load_workbook("gallery.xlsx")
     sheet = excel["Лист1"]
     sheet.append([url, description, title])
     excel.save("gallery.xlsx")

     return render_template("form.html")